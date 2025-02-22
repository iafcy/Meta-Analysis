from adjust_prediction import grubbs_test_outliers, bootlier_plot_outliers, empty_values
from helper import extract_test_results, test_pmid_list
from stat_result import stat_test
import pandas as pd
import numpy as np
import openpyxl
import json
import os

def fixed_effects_pooled_estimate(effect_sizes, weights):
    effect_sizes = np.array(effect_sizes)
    weights = np.array(weights)
    sum_weights = np.sum(weights)
    pooled_estimate = np.sum(weights * effect_sizes) / sum_weights
    return pooled_estimate

def get_path_list(result_type):
    if result_type == "prediction":
        base_dir = "./prediction_results"
    elif result_type == "usage":
        base_dir = "./usage_results"

    list_of_paths = os.listdir(base_dir)
    update_list = []

    for file_path in list_of_paths:
        if result_type != "evaluation":
            file_path = os.path.join(base_dir, file_path)
            update_list.append(file_path)
        else :
            file_path = os.path.join("./evaluate_results", f"{file_path[21:-5]}_evaluation.json")
            update_list.append(file_path)

    return update_list

def get_pred_and_gt(file_path, mod):
    data, empty_num = evaluate_each_meta(file_path, type="stat", mod=mod)

    ground_truth = [point[0] for point in data]
    prediction = [point[1] for point in data]
    return ground_truth, prediction, empty_num

# Evaluate performance of pdf_parser + model combination for each ma
def evaluate_each_meta(file_path, type="eval", mod="overall"):
    with open(file_path, 'r', encoding='utf-8') as json_file:
        data = json.load(json_file)

    results = []
    pooled_estimate_pairs = []
    empty_num = 0

    for meta in data:
        total_number = 0
        sum_abs_diff = 0.0            
        sum_percentage_diff = 0.0     
        table_num = 0
        all_truths = []
        all_preds = []
        
        forest_plots = meta.get('forest_plot_title', [])
        pmid = str(meta.get('pmid', []))
        for forest_plot in forest_plots:
            table_num += 1
            prediction = forest_plot.get('Predicted_ratio', [])
            ground_truth = forest_plot.get('GT_ratio', [])
            weights = [1] * len(ground_truth)

            outlier_index = empty_values(prediction)
            empty_num += len(outlier_index)
            prediction_clean = []
            ground_truth_clean = []
            weights_clean = []

            for idx, val in enumerate(prediction):
                if idx in outlier_index:
                    continue
                try:
                    num = float(val)
                    prediction_clean.append(num)
                    ground_truth_clean.append(ground_truth[idx])
                    weights_clean.append(weights[idx])
                except ValueError:
                    print(f"skip val: {val} (index {idx})")
                    continue
            
            # outlier_index = grubbs_test_outliers(prediction_clean)
            outlier_index = []
            prediction = [float(val) for idx, val in enumerate(prediction_clean) if idx not in outlier_index]
            ground_truth = [val for idx, val in enumerate(ground_truth_clean) if idx not in outlier_index]
            weights = [val for idx, val in enumerate(weights_clean) if idx not in outlier_index]

            df = pd.DataFrame({
                'Ground_Truth': ground_truth,
                'Predicted': prediction,
                'Weights': weights
            })

            truth_pooled_estimate = fixed_effects_pooled_estimate(df['Ground_Truth'], df['Weights'])
            pred_pooled_estimate = fixed_effects_pooled_estimate(df['Predicted'], df['Weights'])

            if mod == "test" and str(pmid) in test_pmid_list:
                pooled_estimate_pairs.append((truth_pooled_estimate, pred_pooled_estimate))

            abs_diff = abs(truth_pooled_estimate - pred_pooled_estimate)
            relative_diff = abs_diff / (abs(truth_pooled_estimate) + 1e-6)

            sum_abs_diff += abs_diff * len(weights)
            sum_percentage_diff += relative_diff * len(weights)

            total_number += len(weights)

            truths = ground_truth_clean
            preds = prediction_clean
            all_truths.extend(truths)
            all_preds.extend(preds)

        final_abs_diff = sum_abs_diff / total_number if total_number > 0 else 0.0
        final_percentage_diff = sum_percentage_diff / total_number if total_number > 0 else 0.0

        # t_paired_test
        t_stat, t_pvalue, ks_stat, ks_pvalue = stat_test(all_truths, all_preds)

        # calculate variance
        variance = sum([(truth - sum(all_truths) / len(all_truths))**2 for truth in all_truths]) / len(all_truths)

        result = {
            "pmid": meta.get('pmid', ""),
            "total_number": total_number,
            "absolute_diff": final_abs_diff,
            "percentage_diff": final_percentage_diff,
            "GT_variance": variance,
            "t_paired_test": {
                "t_stat": t_stat,
                "p_value": t_pvalue
            },
            "ks_test": {
                "ks_stat": ks_stat,
                "p_value": ks_pvalue
            }
        }

        results.append(result)

    if type == "stat":
        return pooled_estimate_pairs, empty_num
    else:
        if not os.path.exists("./metrics_results"):
            os.makedirs("./metrics_results")

        with open(f"./metrics_results/{file_path[21:-5]}_evaluation.json", "w", encoding="utf-8") as json_file:
            json.dump(results, json_file, ensure_ascii=False, indent=4)

# Evaluate the absolute difference and percentage difference of all meta
def evaluate_abs_and_precent_diff(file_path, mod="overall"):
    with open(file_path, "r", encoding="utf-8") as json_file:
        results = json.load(json_file)

    # Sort the entries by 'percentage_diff' in descending order
    sorted_results = sorted(results, key=lambda x: x["percentage_diff"], reverse=True)

    # Remove the top 5 outliers
    cleaned_results = sorted_results[5:]
    test_results = extract_test_results(results)

    def calculate_weighted_stats(results, key):
        total_weight = 0
        weighted_sum = 0.0
        
        scaled_values = []
        for result in results:
            val = result[key]

            scaled_values.append({
                "weight": result["total_number"],
                "value": val
            })

        for item in scaled_values:
            w = item["weight"]
            v = item["value"]
            total_weight += w
            weighted_sum += v * w

        weighted_avg = weighted_sum / total_weight if total_weight > 0 else 0.0

        variance_numerator = 0.0
        for item in scaled_values:
            w = item["weight"]
            v = item["value"]
            variance_numerator += w * ((v - weighted_avg) ** 2)

        variance = variance_numerator / total_weight if total_weight > 0 else 0.0

        # if key == "percentage_diff":
        #     weighted_avg *= 100
        #     variance *= 10000   # 

        return weighted_avg, variance
    
    if mod == "overall":
        avg_absolute_diff, var_absolute_diff = calculate_weighted_stats(cleaned_results, "absolute_diff")
        avg_percentage_diff, var_percentage_diff = calculate_weighted_stats(cleaned_results, "percentage_diff")
    elif mod == "test":
        avg_absolute_diff, var_absolute_diff = calculate_weighted_stats(test_results, "absolute_diff")
        avg_percentage_diff, var_percentage_diff = calculate_weighted_stats(test_results, "percentage_diff")

    # print(f"Absolute Diff - Weighted Average: {avg_absolute_diff:.5f}, Variance: {var_absolute_diff:.5f}")
    # print(f"Percentage Diff - Weighted Average: {avg_percentage_diff * 100:.5f}%, Variance: {var_percentage_diff * 100:.5f}%")

    return [avg_absolute_diff, var_absolute_diff, avg_percentage_diff, var_percentage_diff]

# Include all the metrics and saved to excel
def evaluate_overall_performance(mod):
    path_list = get_path_list("prediction")

    workbook = openpyxl.Workbook()
    sheet = workbook.active  

    headers = [
        "Combinations",  
        "Avg_Abs_Diff",
        "Var_Abs_Diff",
        "Avg_Percentage_Diff",
        "Var_Percentage_Diff",
        "T_Stat",
        "T_Pvalue",
        "KS_Stat",
        "KS_Pvalue",
        "Empty_Response",
        "Empty_Percentage"
    ]

    sheet.append(headers)

    path_list = sorted(path_list)

    for file_path in path_list:

        evaluate_each_meta(file_path, type="eval")

        ground_truth, prediction, empty_num = get_pred_and_gt(file_path, mod=mod)
        all_find_ratio_tasks = 1326
        empty_percentage = empty_num/all_find_ratio_tasks

        file_path = os.path.join("./evaluate_results", f"{file_path[21:-5]}_evaluation.json")
        performance = evaluate_abs_and_precent_diff(file_path, mod=mod)  # [avg_absolute_diff, var_absolute_diff, avg_percentage_diff, var_percentage_diff]
        # avg_prediction = np.mean(prediction)
        # avg_ground_truth = np.mean(ground_truth)

        stat_results = stat_test(ground_truth, prediction)  # [t_stat, t_pvalue, ks_stat, ks_pvalue]

        sheet.append([
            file_path[19:-16],  
            performance[0],    # Avg_Abs_Diff
            performance[1],    # Var_Abs_Diff
            performance[2],    # Avg_Percentage_Diff
            performance[3],    # Var_Percentage_Diff
            stat_results[0],   # T_Stat
            stat_results[1],   # T_Pvalue
            stat_results[2],   # KS_Stat
            stat_results[3],   # KS_Pvalue
            empty_num,
            empty_percentage*100
        ])
    
    if not os.path.exists("./overall_results"):
        os.makedirs("./overall_results")

    excel_file_path = "./overall_results/overall_prediction_results.xlsx"
    workbook.save(excel_file_path)
    print(f"Metrics Results Saved: {excel_file_path}")

# Evaluate the usage of time and money
def evaluate_usages():
    path_list = get_path_list("usage")
    path_list.sort()

    workbook = openpyxl.Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    
    overall_data = {}

    for file_path in path_list:
        filename = os.path.basename(file_path)[:-5]  
        parts = filename.split("_", 1)  
        if len(parts) != 2:
            print(f"file name incorrect: {filename}")
            continue
            
        parser_type, model = parts
        combo_name = f"{parser_type}_{model}"
        
        if combo_name not in overall_data:
            overall_data[combo_name] = {
                "time_total": 0.0,
                "token_total": 0,
                "find_ratio_total": 0.0,
                "cost_total": 0.0,
                "count": 0
            }

        sheet = workbook.create_sheet(title=combo_name)
        sheet.append(["PMID", "Time_Used", "Token_Used", "Find_Ratio_Tasks", "Cost"])

        with open(file_path, "r", encoding="utf-8") as json_file:
            data = json.load(json_file)

        model_price = 0
        one_million = 1_000_000
        if "Claude-3.5" in model:
            model_price = 3 / one_million
        elif "4o-mini" in model:
            model_price = 0.15 / one_million
        elif "4o" in model:
            model_price = 2.5 / one_million

        llamaParse_price = 3 / 1000

        for meta in data:
            time_used = meta.get("time_used", 0)
            token_used = meta.get("token_used", 0)
            find_ratio = meta.get("find_ratio_tasks", 0)
            
            if parser_type == "llamaParser":
                pdf_pages = meta.get("pdf_pages", 0)
                cost = token_used * model_price + pdf_pages * llamaParse_price
            else:
                cost = token_used * model_price

            sheet.append([
                meta.get("pmid", ""),
                time_used,
                token_used,
                find_ratio,
                cost
            ])

            overall_data[combo_name]["time_total"] += time_used
            overall_data[combo_name]["token_total"] += token_used
            overall_data[combo_name]["find_ratio_total"] += find_ratio
            overall_data[combo_name]["cost_total"] += cost
            overall_data[combo_name]["count"] += 1

    overall_sheet = workbook.create_sheet(title="Overall")

    headers = ["Combination", "Average Time", "Average Tokens", "Average Find Ratio", "Total Cost"]
    overall_sheet.append(headers)

    for combo_name, metrics in overall_data.items():
        if metrics["count"] == 0:
            continue
        avg_time = metrics["time_total"] / metrics["count"]
        avg_tokens = metrics["token_total"] / metrics["count"]
        avg_find_ratio = metrics["find_ratio_total"] / metrics["count"]
        total_cost = metrics["cost_total"]
        
        overall_sheet.append([combo_name, avg_time, avg_tokens, avg_find_ratio, total_cost])

    if not os.path.exists("./overall_results"):
        os.makedirs("./overall_results")

    excel_file_path = "./overall_results/usage_results.xlsx"
    workbook.save(excel_file_path)
    print(f"Usage Results Saved: {excel_file_path}")

def evaluate_find_ratio():
    evaluate_each_meta()

    # Saved Results to excel
    evaluate_overall_performance()
    evaluate_usages()
    print("Finished all evaluations for find_ratio")






